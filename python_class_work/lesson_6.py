from openpyxl import load_workbook
import requests

# 读取数据函数
def read_data(file_name, sheet_name):
    work_book = load_workbook(file_name)
    sheet = work_book[sheet_name]
    # cell = sheet.cell(row=1, column=1)
    # print(cell.value)
    # cell.value = "张未久"
    # work_book.save("test_case_api.xlsx")
    cases = []
    max_row = sheet.max_row
    for i in range (2,max_row+1):
        dict_case = dict(
        case_id=sheet.cell(row=i, column=1).value,
        url=sheet.cell(row=i , column=5).value,
        data=sheet.cell(row=i, column=6).value,
        expected_result=sheet.cell(row=i, column=7).value)
        cases.append(dict_case)
    return cases
result_reg = read_data("test_case_api.xlsx", "register")
result_log = read_data("test_case_api.xlsx", "login")
print(result_reg)
print(result_log)

# 写入数据函数
def write_result(file_name, sheet_name, final_result, row, column):
    work_book = load_workbook(file_name)
    sheet = work_book[sheet_name]
    sheet.cell(row=row, column=column).value = final_result
    work_book.save(file_name)

# 发送接口请求函数
def api_func(url_api, data_api):
    head = {"X-Lemonban-Media-Type": "lemonban.v2"}
    response = requests.post(url=url_api, json=data_api, headers=head)
    return response.json()

if __name__ == '__main_':
    re = read_data("test_case_api.xlsx", "register")
    print(re)





